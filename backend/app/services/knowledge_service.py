import hashlib
import json
import math
import re
from datetime import datetime
from pathlib import Path

from sqlalchemy import text

from app.services.database import engine


KNOWLEDGE_CATEGORIES = (
    "Penglukatan",
    "Pembayuhan",
    "Tenung",
    "Permata",
    "Lontar",
    "Wariga",
    "Pengetahuan Lain",
)
EMBEDDING_DIMENSIONS = 128
MAX_CHUNK_CHARACTERS = 1200
CHUNK_OVERLAP_CHARACTERS = 180
TOKEN_PATTERN = re.compile(r"[A-Za-zÀ-ÖØ-öø-ÿ0-9_]+")


def ensure_knowledge_tables(conn):
    conn.execute(
        text("""
            CREATE TABLE IF NOT EXISTS knowledge_documents (
                id SERIAL PRIMARY KEY,
                title TEXT NOT NULL,
                category TEXT NOT NULL,
                source_filename TEXT NOT NULL,
                file_type TEXT NOT NULL,
                raw_text TEXT NOT NULL,
                created_at TIMESTAMP NOT NULL DEFAULT NOW()
            )
        """)
    )
    conn.execute(
        text("""
            CREATE TABLE IF NOT EXISTS knowledge_chunks (
                id SERIAL PRIMARY KEY,
                document_id INTEGER NOT NULL REFERENCES knowledge_documents(id)
                    ON DELETE CASCADE,
                category TEXT NOT NULL,
                chunk_index INTEGER NOT NULL,
                content TEXT NOT NULL,
                embedding TEXT NOT NULL,
                created_at TIMESTAMP NOT NULL DEFAULT NOW()
            )
        """)
    )


def normalize_category(category):
    normalized = str(category or "").strip()

    if normalized in KNOWLEDGE_CATEGORIES:
        return normalized

    raise ValueError("Kategori pengetahuan tidak valid.")


def normalize_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def tokenize(text_value):
    return [
        token.lower()
        for token in TOKEN_PATTERN.findall(text_value)
        if len(token) > 2
    ]


def build_embedding(text_value):
    vector = [0.0] * EMBEDDING_DIMENSIONS

    for token in tokenize(text_value):
        digest = hashlib.md5(token.encode("utf-8")).digest()
        index = int.from_bytes(digest[:4], "big") % EMBEDDING_DIMENSIONS
        sign = 1 if digest[4] % 2 == 0 else -1
        vector[index] += sign

    norm = math.sqrt(sum(value * value for value in vector))

    if norm == 0:
        return vector

    return [round(value / norm, 6) for value in vector]


def cosine_similarity(left, right):
    return sum(a * b for a, b in zip(left, right))


def chunk_text(raw_text):
    cleaned_text = normalize_text(raw_text)

    if not cleaned_text:
        return []

    chunks = []
    start = 0

    while start < len(cleaned_text):
        end = min(start + MAX_CHUNK_CHARACTERS, len(cleaned_text))

        if end < len(cleaned_text):
            sentence_end = max(
                cleaned_text.rfind(". ", start, end),
                cleaned_text.rfind("? ", start, end),
                cleaned_text.rfind("! ", start, end),
                cleaned_text.rfind("\n", start, end),
            )

            if sentence_end > start + 400:
                end = sentence_end + 1

        chunk = cleaned_text[start:end].strip()

        if chunk:
            chunks.append(chunk)

        if end >= len(cleaned_text):
            break

        start = max(0, end - CHUNK_OVERLAP_CHARACTERS)

    return chunks


def extract_pdf_text(file_bytes):
    try:
        from pypdf import PdfReader
    except ImportError as error:
        raise ValueError(
            "Upload PDF membutuhkan dependency pypdf. Jalankan pip install -r requirements.txt."
        ) from error

    reader = PdfReader(file_bytes)
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n".join(pages)


def extract_excel_text(file_bytes):
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ValueError(
            "Upload Excel membutuhkan dependency openpyxl. Jalankan pip install -r requirements.txt."
        ) from error

    workbook = load_workbook(file_bytes, read_only=True, data_only=True)
    lines = []

    for sheet in workbook.worksheets:
        lines.append(f"Sheet: {sheet.title}")

        for row in sheet.iter_rows(values_only=True):
            values = [normalize_text(value) for value in row if normalize_text(value)]

            if values:
                lines.append(" | ".join(values))

    return "\n".join(lines)


def extract_text_from_upload(file_bytes, filename):
    suffix = Path(filename or "").suffix.lower()

    if suffix == ".pdf":
        return extract_pdf_text(file_bytes)

    if suffix in (".xlsx", ".xlsm"):
        return extract_excel_text(file_bytes)

    if suffix in (".txt", ".md", ".csv"):
        return file_bytes.read().decode("utf-8", errors="replace")

    raise ValueError("Format file harus PDF, Excel (.xlsx/.xlsm), atau Text (.txt/.md/.csv).")


def create_knowledge_document(category, title, filename, file_bytes):
    normalized_category = normalize_category(category)
    normalized_title = normalize_text(title) or Path(filename).stem
    raw_text = normalize_text(extract_text_from_upload(file_bytes, filename))
    chunks = chunk_text(raw_text)

    if not raw_text or not chunks:
        raise ValueError("File tidak memiliki teks yang bisa disimpan sebagai knowledge.")

    with engine.begin() as conn:
        ensure_knowledge_tables(conn)
        document_id = conn.execute(
            text("""
                INSERT INTO knowledge_documents
                    (title, category, source_filename, file_type, raw_text)
                VALUES
                    (:title, :category, :source_filename, :file_type, :raw_text)
                RETURNING id
            """),
            {
                "title": normalized_title,
                "category": normalized_category,
                "source_filename": filename,
                "file_type": Path(filename).suffix.lower().lstrip(".") or "text",
                "raw_text": raw_text,
            },
        ).scalar_one()

        for index, chunk in enumerate(chunks):
            conn.execute(
                text("""
                    INSERT INTO knowledge_chunks
                        (document_id, category, chunk_index, content, embedding)
                    VALUES
                        (:document_id, :category, :chunk_index, :content, :embedding)
                """),
                {
                    "document_id": document_id,
                    "category": normalized_category,
                    "chunk_index": index,
                    "content": chunk,
                    "embedding": json.dumps(build_embedding(chunk)),
                },
            )

    return {
        "id": document_id,
        "title": normalized_title,
        "category": normalized_category,
        "source_filename": filename,
        "chunk_count": len(chunks),
        "created_at": datetime.utcnow().isoformat(),
    }


def list_knowledge_documents():
    with engine.begin() as conn:
        ensure_knowledge_tables(conn)
        rows = conn.execute(
            text("""
                SELECT
                    d.id,
                    d.title,
                    d.category,
                    d.source_filename,
                    d.file_type,
                    d.created_at,
                    COUNT(c.id) AS chunk_count
                FROM knowledge_documents d
                LEFT JOIN knowledge_chunks c ON c.document_id = d.id
                GROUP BY d.id
                ORDER BY d.created_at DESC, d.id DESC
            """)
        )

        return [dict(row._mapping) for row in rows]


def delete_knowledge_document(document_id):
    try:
        normalized_id = int(document_id)
    except (TypeError, ValueError) as error:
        raise ValueError("ID knowledge tidak valid.") from error

    with engine.begin() as conn:
        ensure_knowledge_tables(conn)
        document = conn.execute(
            text("""
                SELECT
                    id,
                    title,
                    category,
                    source_filename
                FROM knowledge_documents
                WHERE id = :document_id
            """),
            {"document_id": normalized_id},
        ).mappings().first()

        if not document:
            raise ValueError("Knowledge tidak ditemukan.")

        deleted_chunks = conn.execute(
            text("""
                DELETE FROM knowledge_chunks
                WHERE document_id = :document_id
            """),
            {"document_id": normalized_id},
        ).rowcount

        conn.execute(
            text("""
                DELETE FROM knowledge_documents
                WHERE id = :document_id
            """),
            {"document_id": normalized_id},
        )

    return {
        "id": document["id"],
        "title": document["title"],
        "category": document["category"],
        "source_filename": document["source_filename"],
        "deleted_chunk_count": deleted_chunks,
    }


def search_knowledge(query, limit=5):
    normalized_query = normalize_text(query)

    if not normalized_query:
        return []

    query_embedding = build_embedding(normalized_query)
    query_tokens = set(tokenize(normalized_query))

    with engine.begin() as conn:
        ensure_knowledge_tables(conn)
        rows = conn.execute(
            text("""
                SELECT
                    c.content,
                    c.embedding,
                    c.category,
                    d.title,
                    d.source_filename
                FROM knowledge_chunks c
                JOIN knowledge_documents d ON d.id = c.document_id
                ORDER BY c.created_at DESC
                LIMIT 800
            """)
        )

        scored_chunks = []

        for row in rows:
            item = dict(row._mapping)
            chunk_tokens = set(tokenize(item["content"]))
            keyword_score = len(query_tokens & chunk_tokens) / max(len(query_tokens), 1)
            vector_score = cosine_similarity(query_embedding, json.loads(item["embedding"]))
            score = vector_score + keyword_score

            if score <= 0:
                continue

            scored_chunks.append({**item, "score": score})

    scored_chunks.sort(key=lambda item: item["score"], reverse=True)
    return scored_chunks[:limit]


def build_knowledge_context(query):
    results = search_knowledge(query)

    if not results:
        return "Tidak ada knowledge upload yang relevan ditemukan di database."

    context_blocks = []

    for index, item in enumerate(results, start=1):
        context_blocks.append(
            "\n".join(
                [
                    f"[Knowledge {index}]",
                    f"Kategori: {item['category']}",
                    f"Judul: {item['title']}",
                    f"Sumber: {item['source_filename']}",
                    f"Isi: {item['content']}",
                ]
            )
        )

    return (
        "Gunakan knowledge upload berikut sebagai sumber saat relevan. "
        "Jika konteks tidak cukup, katakan bahwa data belum tersedia dan jangan mengarang.\n"
        + "\n\n".join(context_blocks)
    )
